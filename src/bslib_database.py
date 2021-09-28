"""This files contains function to load and edit the data base for bslib"""
import pandas as pd
from openpyxl import load_workbook


def load_parameter(fname):
    """Loads system parameter from excel file

    :param fname: Path to the excel file
    :type fname: string
    :param col_name: Column to read data from
    :type col_name: string
    :return: Dictionary holding parameters from the Excel sheet
    :rtype: dict
    """

    wb = load_workbook(fname, data_only=True)
    ws = wb['Data']  # Load Data sheet of excel file

    # read keys and values from Excel sheet
    keys = (c.value for c in ws['E'][1:])

    for col_name in ws.iter_cols(min_col=7):

        values = [c.value if c.value != 'ns' else None for c in ws[col_name][1:]]

        if "End" in values:
            break

        parameter = dict(zip(keys, values))

        # deletes entries where key is None
        del parameter[None]

        # Assign specific parameters
        parameter['P_PV2AC_out_PVINV'] = ws[col_name][15].value
        parameter['P_PV2AC_out'] = ws[col_name][24].value
        parameter['P_AC2BAT_in_DCC'] = ws[col_name][25].value
        parameter['P_AC2BAT_in'] = ws[col_name][26].value
        parameter['P_BAT2AC_out'] = ws[col_name][27].value
        parameter['P_BAT2AC_out_DCC'] = ws[col_name][28].value

        # Specific parameters of DC-coupled systems
        if parameter['Top'] == 'DC':
            parameter['P_AC2BAT_in'] = parameter['P_AC2BAT_in_DCC']  # Nominal charging power (AC) in kW
            parameter['P_BAT2AC_out'] = parameter['P_BAT2AC_out_DCC']

        # Specific parameters of PV inverters and AC-coupled systems
        if parameter['Top'] == 'PVINV' or parameter['Top'] == 'AC' and parameter['P_PV2AC_out_PVINV'] is not None:
            parameter['P_PV2AC_out'] = parameter['P_PV2AC_out_PVINV']

        # Specific parameters of PV-coupled systems
        if parameter['Top'] == 'PV':
            parameter['P_BAT2PV_in'] = parameter['P_BAT2AC_in']
            parameter['P_BAT2AC_out'] = parameter['P_BAT2AC_out_DCC']

        # replace 'ns', 'o' and 'c' entries to None
        for key, value in parameter.items():
            if value == 'ns' or value == 'o' or value == 'c' or value == ' ':
                parameter[key] = None

        # Convert to kW
        convert_to_kw = ['P_PV2AC_in', 'P_PV2AC_out_PVINV', 'P_PV2AC_out', 'P_AC2BAT_in_DCC', 'P_AC2BAT_in', 'P_BAT2AC_out',
                         'P_BAT2AC_out_DCC', 'P_PV2BAT_in', 'P_BAT2PV_out', 'P_PV2BAT_out', 'P_BAT2AC_in']

        for par in convert_to_kw:
            if parameter[par]:
                parameter[par] /= 1000

    return parameter


def transpose_df(df):
    df = df.T
    return df


def main():
   parameter = load_parameter("/Users/kairosken/Documents/HS-Emden/Arbeit/bslib/input/PerModPAR.xlsx")


def export_to_csv(df):
    df.to_csv('../input/out.csv', index=False, header=None)


if __name__ == "__main__":
    main()
